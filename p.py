from openpyxl import load_workbook
import json

planilha = load_workbook(filename='Digitação Inovação Completo com CNAE.xlsx')
planilha_banco_de_dados = load_workbook(filename='SP.xlsx')


digitacao = planilha['Digitação']
banco_de_dados = planilha_banco_de_dados['Arquivo']



class Excel:
    def __init__(self, p, v1=str, v2=str):
        self.campo = p[v1:v2]
        self.lista = {}
        self.lista_iguais = {}
        self.contador = {}
       
        print('Campo adicionado')
        self.adiciona_valores()

        
    def busca_duplicatas(self, i=dict):

       
    
        for l,v in self.lista.items():
            for l1, v1 in i.items():
                if v == v1:
                    print(f'{l1} {v1} Duplicado')
                    print('{v1} Adicionado na lista')
                    self.contador[l1] = v1
                    self.lista_iguais[l1] = v1
                  
           
             
                else:
                    print(l,'<= Linha planilha |', 'Duplicatas encontradas =>', len(self.contador))
        
        json_object = json.dumps(self.lista_iguais, indent=4)
        with open('telefones', 'w') as arquivo_saida:
            arquivo_saida.write(json_object)
            print('Arquivo finalizado')





    def adiciona_valores(self):
        for i in self.campo:
            self.lista[i.row] = str(i.value)






    





tel1_digitacao = Excel(digitacao, 'e', 'e')


tel1_banco = Excel(banco_de_dados, 'l', 'l')







tel1_digitacao.busca_duplicatas(tel1_banco.lista)







        